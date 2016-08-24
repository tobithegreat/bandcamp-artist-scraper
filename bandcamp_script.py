'''This will be a python script designed to glean the info for the artists on the Bandcamp page. The script will involve using the Internet, and modules associated will include Selenium, and the Excel Module. This also requires the use of the Firefox browser. '''

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

#==============================================================================
# from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
#==============================================================================
import openpyxl
#==============================================================================
# # Used to control browser programmatically, clicking on artist page links and going back to directory page.
#==============================================================================


### Open the browser to Firefox, and directs to the Bandcamp Discover Directory, with the New Arrivals filter selected.
#browser = webdriver.Firefox()

# FIX: Newer Firefox versions are not compatible with Selenium. Download an older firefox version (I used 35.0), and change path below to the new binary
# Change the path below to your Firefox directory (older version)
#==============================================================================
# browser = webdriver.Firefox(firefox_binary=FirefoxBinary("/Volumes/Firefox/Firefox.app/Contents/MacOS/firefox-bin"))
#==============================================================================
browser = webdriver.Chrome() 
browser.get('https://bandcamp.com/?g=all&s=new&p=0&gn=0&f=all&w=0')

browser.maximize_window()

#==============================================================================
# browser.find_element_by_class_name("discover-link").click()
#==============================================================================

browser.execute_script("window.scrollTo(0, 250)")

book = openpyxl.load_workbook(filename = "Bandcamp Artists.xlsx")
sheet = book.worksheets[0]

## Selects the <a> tag that corresponds to the artist preview page, that will open on the right side of the screen, and clicks it.

## For each artist...

def go_through_page():
  
    for div_index in range(0, 8):
        

        current_section = browser.find_element_by_css_selector('div.row.discover-result.result-current')
        artist_section_1 = current_section.find_elements_by_css_selector("div.col.col-3-12.discover-item")
     
        print("Page ", i," ARTIST ",div_index + 1)
        
        artist_section_1 = artist_section_1[div_index]
        
        artist_genre = artist_section_1.find_element_by_class_name("item-genre").text
   
        print("Type : ", artist_genre)
        
        
#==============================================================================
#         browser.execute_script("window.scrollTo(0, document.body.scrollHeight/6.5);")
#         browser.maximize_window()
#==============================================================================
#==============================================================================
#         element=browser.find_element_by_class_name("section-title")
#         element.location_once_scrolled_into_view
#==============================================================================
        
        artist_link = artist_section_1.find_element_by_tag_name('a')
        artist_link.location_once_scrolled_into_view
#==============================================================================
#         WebDriverWait(browser, 10).until(EC.element_to_be_clickable(artist_section_1.find_element_by_tag_name('a')))
#==============================================================================
        
        artist_link.click()

## From the sub-menu that pops up on the right, selects the sub-element of the <p> tag, then clicks it. This sub-element will be the <a> tag that leads to the artist page.
        artist_page = browser.find_element_by_xpath("//p[@class = 'detail-artist']/a")
        artist_page.location_once_scrolled_into_view
        artist_page.click()


        artist_section_2 = browser.find_element_by_xpath("//p[@id = 'band-name-location']")
        artist_name = artist_section_2.find_element_by_class_name('title').text
        print("Name : ",artist_name)
        artist_bandcamp_url = browser.current_url
        print("URL : ", artist_bandcamp_url)
        try:
            artist_location = artist_section_2.find_element_by_class_name("location").text
        except:
            artist_location = None
            
        print ("Location : ", artist_location)
        artist_email = None
        artist_twitter = None
        try:
            artist_links = browser.find_element_by_xpath("//ol[@id = 'band-links']")
            artist_facebook = artist_links.find_element_by_partial_link_text("Facebook").get_attribute('href')
            
            
        except:
            artist_facebook = None
        
        
        print("Facebook : ", artist_facebook)
        currentRow = div_index + 8*(i - 1) + 2
        write_to_sheet(sheet, currentRow , artist_name, artist_bandcamp_url, artist_location, \
        artist_genre, artist_email, artist_twitter, artist_facebook)
        
        browser.back()
        WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR,"div.col.col-3-12.discover-item")))
        
        
## TODO: Write to Excel Spreadsheet Row, and move to next Row. Then, continue in for loop for all the artists on the page. Should be 8 per page. 
#==============================================================================
#         
#==============================================================================


## Move to the next page.
def next_directory_page():
    browser.get('https://bandcamp.com/?g=all&s=new&p='+str(i-1) +'&gn=0&f=all&w=0')
#==============================================================================
#     button_section = browser.find_element_by_xpath("//div[@class = 'pages']")
#     next_button = button_section.find_element_by_link_text("next")
#     next_button.location_once_scrolled_into_view
#     next_button.click()
#==============================================================================

    WebDriverWait(browser, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR,'div.row.discover-result.result-current')))
   

### For each page:
##Go into each artist's page.
## Grab the Information from each artist.
## Write the information from each artist into a row in the spreadsheet)


## A method to add artist info to a row in the spreadsheet, and increment the row value for the next row.
# FIX: Default argument was in the middle of non-default ones. Moved it back to correct position
def write_to_sheet(sheet, row_Value, artist, url, location, genre , email, twitter, facebook):
    sheet.cell(row = row_Value, column = 1).value = artist
    sheet.cell(row = row_Value, column = 2).value = url
    sheet.cell(row = row_Value, column = 3).value = genre
    sheet.cell(row = row_Value, column = 4).value = location
    sheet.cell(row = row_Value, column = 5).value = email
    sheet.cell(row = row_Value, column = 6).value = twitter
    sheet.cell(row = row_Value, column = 7).value = facebook
   
    
## Go back to the directory and go to the next page.

write_to_sheet(sheet, 1, 'Artist', 'URL', 'Location', 'Genre' , 'Email', 'Twitter', 'Facebook')
for i in range (1, 201):
    
    go_through_page()
    next_directory_page()
    book.save("Bandcamp Artists.xlsx")
  
    
   


